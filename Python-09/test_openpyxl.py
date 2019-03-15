import openpyxl

# 加载工作薄
wb = openpyxl.load_workbook("2.xlsx")
print(type(wb))
print(wb.sheetnames)
#获取工作薄中的激活的sheet
sheet = wb.active
# 获取sheet的标题
print(sheet.title)
# 修改sheet的标题 
# sheet.title = "newsheet"
# wb.create_sheet("201903")

# 住sheet中的单元格写数据
# sheet.cell(1,1).value = "日期"
# print(sheet.cell(1,1).value)
# sheet['B1'] = "姓名"

rows = sheet.rows
rd = sheet.row_dimensions
# print(rows)
print(rd)
# for row in rows:
# 	print(row)
print(sheet.max_row)
print(sheet.max_column)
wb.remove_sheet()
# 保存
wb.save("2.xlsx")
